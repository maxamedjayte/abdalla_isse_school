
from email import header
from tkinter.tix import Tree
from wsgiref.util import FileWrapper
from attr import field
from django.http import HttpResponse, response
from django.shortcuts import render,redirect
from django.http.response import HttpResponseRedirect, JsonResponse
from django.contrib.auth import authenticate,login,logout
from django.contrib.auth.decorators import login_required

from rest_framework.decorators import api_view
from rest_framework.response import Response
from sqlalchemy import true
from api import serializers

from api.models import *
from api.serializers import ClasseSerializer, StudentSerializer
from openpyxl import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import os
import openpyxl
from openpyxl.drawing.image import Image

# Create your views here.
def adminLoginPage(request):
    status=''
    if request.POST:
        username=request.POST.get('username')
        password=request.POST.get('password')

        if Student.objects.filter(studentId=username).exists():
            status='wrong admin username'
            data={
                'status':status
            }
            return render(request,'normal_pages/login.html',data)
        else:
            user = authenticate(request,username=username,password=password)
            if user is not None:
                login(request,user)
                return redirect('/school-admin/')
            else:
                status='Usernameka ama Passwordka ayaa qaldan'
    data={
        'status':status
    }
    return render(request,'normal_pages/login.html',data)

def adminLogoutPage(request):
    logout(request)
    return HttpResponseRedirect('/adminlogin/')



# zfront-website
def indexWeb(reqesut):
    return render(reqesut,'zfront-website/index.html')


@login_required(login_url='/adminlogin/')
def dashboard(request):
    tiradaWiilasha=Student.objects.filter(gender='wiil').count()
    tiradaGabdhaha=Student.objects.filter(gender='gabar').count()
    tiradaFasalada=Classe.objects.filter().count()
    if Student.objects.filter(studentId=request.user.username).exists():
        userId=Student.objects.get(studentId=request.user.username).pk
        return  HttpResponseRedirect('/adminlogin/')
    else:
        return render(request,'normal_pages/dashboard.html',{'tiradaGabdhaha':tiradaGabdhaha,'tiradaWiilasha':tiradaWiilasha,'tiradaFasalada':tiradaFasalada})

@login_required(login_url='/adminlogin/')
def studentRegManag(request):
    if Student.objects.filter(studentId=request.user.username).exists():
        userId=Student.objects.get(studentId=request.user.username).pk
        return  HttpResponseRedirect('/adminlogin/')
    else:
        return render(request,'reg_managing/students.html')

@login_required(login_url='/adminlogin/')
def xogtaArdayda(request):
    return render(request,'managing/xogta_students.html')


@login_required(login_url='/adminlogin/')
def thisStdInfo(request,pk):
    if Student.objects.filter(studentId=request.user.username).exists():
        userId=Student.objects.get(studentId=request.user.username).pk
        return  HttpResponseRedirect('/adminlogin/')
    else:
        return render(request,'detail/student_info.html',{'pk':pk})




# class
@login_required(login_url='/adminlogin/')
def classRegMang(request):
    return render(request,'reg_managing/classes.html')

@login_required(login_url='/adminlogin/')
def xogtaFasalada(request):
    return render(request,'managing/xogta_fasalada.html')


# sanad dugsiyedka

@login_required(login_url='/adminlogin/')
def sanadDugsiyedkaRegMang(request):
    return render(request,'reg_managing/academic_year.html')

# subjects
@login_required(login_url='/adminlogin/')
def subjects(request):
    return render(request,'reg_managing/subjects.html')


# imtixaanada nooca
@login_required(login_url='/adminlogin/')
def examTypeRegMang(request):
    return render(request,'reg_managing/noocyadaImtixaanada.html')



# imtixaanada galida
@login_required(login_url='/adminlogin/')
def imtixanRegMang(request):
    return render(request,'reg_managing/imtixaanada.html')



@login_required(login_url='/adminlogin/')
def examEntringShow(request):
    return render(request,'managing/examEntringClasses.html')



@login_required(login_url='/adminlogin/')
def examEntringStdShow(request,pk):
    return render(request,'managing/examEntringStudents.html',{'pk':pk})


@login_required(login_url='/adminlogin/')
def studentExamMang(request,pk):
    return render(request,'managing/student-exam-portal.html',{'pk':pk})





# invoice

@login_required(login_url='/adminlogin/')
def examClassInvoice(request,pk):
    return render(request,'invoices/class-exam.html',{'pk':pk})



def test(request):
    print(Marks.objects.all().select_related('studentId'))
    return Response()



@login_required(login_url='/adminlogin/')
def assaignEachClassStdCount(request):
    classCount=5
    allClasse=Classe.objects.all()
    allStudents=Student.objects.all()
    for thisClass in allClasse:
        classCount=Student.objects.filter(stdClass=thisClass.id).count()
        thisClass.studentsCount=classCount
        thisClass.save()
    
    return HttpResponse()

@login_required(login_url='/adminlogin/')
def classTransfer(request):
    return render(request,'managing/classTransfering.html')



def examFileUploadtest(request):
      
            
    return HttpResponse('uploaded')

def checkIfThisMarkCreated(examEntring,subjectId,studentId,stdClasse):
    mark=''
    try:
        mark=Marks.objects.filter(studentId=studentId).filter(stdClasse=stdClasse).filter(subjectId=subjectId).get(theExamId=examEntring)
        
        return {'created':True,'data':mark}
    except:
        serializer={'marks':0}
        return {'created':False,'data':mark}
        
@api_view(['POST'])
def examFileUpload(request):
    fileUploadingStatus=False
    if Student.objects.filter(studentId=request.user.username).exists():
        userId=Student.objects.get(studentId=request.user.username).pk
        return  HttpResponseRedirect('/adminlogin/')
    else:
        if request.method=='POST':
            examFilePost = request.FILES.get('file')
            classId =Classe.objects.get(pk=request.POST.get('classId')) 
            userId = User.objects.get(pk=request.POST.get('userId')) 
            excelFile=ExamEntringFiles.objects.create(examEntringFile=examFilePost,theClasse=classId,uploadedUser=userId)      
            classInfo=Classe.objects.get(pk=classId.pk)
            schoolExams=ExamType.objects.all()
            thisYearExamEntring=ExamEntring.objects.filter(academicYear=AcademicYear.objects.get(isCurrentAcademicYear=True))
            students=Student.objects.filter(stdClass=classInfo.pk).all()
            subjectsOfTheClasse=classInfo.subjects.all()
            
            file=load_workbook(excelFile.examEntringFile)
            file.active
            stdExamSheet=file['Sheet']
            rows=stdExamSheet.rows
            
            clmIndex=1
            stRow=8
            studentId =''
            subjectId=''
            examId=''
            marks=0
            for i,thisStd in enumerate(students):
                studentId=stdExamSheet.cell(row=stRow+2+i, column=clmIndex).value
                # print(stdExamSheet.cell(row=stRow+2+i, column=clmIndex+1).value)

                for subject in subjectsOfTheClasse:
                    subjectId=subject.id
                    for examEntring in thisYearExamEntring:
                        marks=stdExamSheet.cell(row=stRow+2+i, column=clmIndex+2).value
                        # print('marks',stdExamSheet.cell(row=stRow+2+i, column=clmIndex+2).value)            
                        clmIndex=clmIndex+1
                        marksState=checkIfThisMarkCreated(examEntring.id,subjectId,Student.objects.get(studentId=studentId),classInfo.pk)
                        if marksState['created']==False:
                            Marks.objects.create(theExamId=ExamEntring.objects.get(pk=examEntring.id),studentId=Student.objects.get(studentId=studentId),subjectId=Subject.objects.get(pk=subjectId),stdClasse=Classe.objects.get(pk=classInfo.id),marks=marks)
                            fileUploadingStatus=True
                        else:
                            updatedMark=Marks.objects.get(pk=marksState['data'].pk)
                            updatedMark.marks=marks
                            updatedMark.save()
                            fileUploadingStatus=True
                            
                            
                stRow=8
                clmIndex=1
            

            return Response({'status':fileUploadingStatus})










def getThisSubjectMarks(examEntring,subjectId,studentId,stdClasse):
    mark=''
    try:
        mark=Marks.objects.filter(studentId=studentId).filter(stdClasse=stdClasse).filter(subjectId=subjectId).get(theExamId=examEntring)
        return {'created':True,'data':mark}
    except:
        mark={'marks':0}
        return {'created':False,'data':mark}


@api_view(['GET'])
def examFileExporting(request,pk):
    classInfo=Classe.objects.get(pk=pk)
    schoolExams=ExamType.objects.all()
    thisYearExamEntring=ExamEntring.objects.filter(academicYear=AcademicYear.objects.get(isCurrentAcademicYear=True))
    students=Student.objects.filter(stdClass=classInfo.pk).all()
    subjectsOfTheClasse=classInfo.subjects.all()
    clmIndex=1
    stRow=8
    wb=Workbook()
    
    stdExamSheet=wb.active
    
    img = openpyxl.drawing.image.Image('static/img/logo-header.png')
    img.width = 500.21
    img.width = 10.21
    stdExamSheet.add_image(img, 'A1')

    stdExamSheet.cell(row=stRow, column=clmIndex).value = 'STD-ID'
    stdExamSheet.cell(row=stRow, column=clmIndex).font=Font(bold=True,size=11)
    stdExamSheet.cell(row=stRow, column=clmIndex+1).value = 'Student-Name'
    stdExamSheet.cell(row=stRow, column=clmIndex+1).font=Font(bold=True,size=11)
    for subject in subjectsOfTheClasse:
        for exam in schoolExams:
            stdExamSheet.cell(row=stRow, column=clmIndex+2).value = f'{subject}'
            stdExamSheet.cell(row=stRow, column=clmIndex+2).font=Font(bold=True,size=10)
            stdExamSheet.cell(row=stRow+1, column=clmIndex+2).value = f'{exam.examName}'
            stdExamSheet.cell(row=stRow+1, column=clmIndex+2).font=Font(size=8)
            clmIndex=clmIndex+1
    clmIndex=1
    for i,thisStd in enumerate(students):
        print(thisStd)
        stdExamSheet.cell(row=stRow+2+i, column=clmIndex).value = f'{thisStd.studentId}'
        stdExamSheet.column_dimensions[get_column_letter(clmIndex)].width =max(
            stdExamSheet.column_dimensions[get_column_letter(clmIndex)].width,len(f'{thisStd.studentId}')
        )
        stdExamSheet.cell(row=stRow+2+i, column=clmIndex+1).value = f'{thisStd.name}'
        stdExamSheet.column_dimensions[get_column_letter(clmIndex+1)].width =max(
            stdExamSheet.column_dimensions[get_column_letter(clmIndex+1)].width,len(f'{thisStd.name}')
        )
       
        for subject in subjectsOfTheClasse:
            for examEntring in thisYearExamEntring:
                theMarkRes=getThisSubjectMarks(examEntring.id,subject.id,thisStd.id,thisStd.stdClass.pk)
                if theMarkRes['created']==False:
                    stdExamSheet.cell(row=stRow+2+i, column=clmIndex+2).value = str(theMarkRes['data']['marks'])
                else:
                    stdExamSheet.cell(row=stRow+2+i, column=clmIndex+2).value = str(theMarkRes['data'].marks)
                clmIndex=clmIndex+1
        stRow=8
        clmIndex=1
    file_name=classInfo.name+'--'+classInfo.classRamaq+' '+str(classInfo.academicYear.fromDate).split('-')[0]+'-'+str(classInfo.academicYear.toDate).split('-')[0]
    wb.save('examExportedFiles/'+file_name+' '+'.xlsx')
    
    downladFile='examExportedFiles/'+file_name+' '+'.xlsx'

    # worksheet.insert_image('B12', 'https://images.ctfassets.net/hrltx12pl8hq/7yQR5uJhwEkRfjwMFJ7bUK/dc52a0913e8ff8b5c276177890eb0129/offset_comp_772626-opt.jpg', {'x_offset': 15, 'y_offset': 10})

    # base_dir=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # file_path=base_dir+'/'+file_name 
    # file_name=os.path.basename(file_path)
    # chunk_size=8192
    # resp=StreamingHttpResponse(FileWrapper(open(file_path,'rb'),chunk_size,
    #     content_type=mimetypes.guess_type(file_path)[0]))
    # resp['Content-lenght']=os.path.getsize()
    # resp['Content-Disposition']="Attachment;filename=%s"%file_name
    direct=os.getcwd()
    return Response({'status':True,'file':str(direct)+'/'+downladFile})



    


def examFiles(request):
    return render(request,'managing/examFiles.html')